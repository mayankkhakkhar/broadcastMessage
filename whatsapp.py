import xlrd
from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager


print('Enter message to broadcast....') 
string = str(input())
message = string.replace(' ', '%20')

loc = ("./whatsapp.xlsx")
wb = openpyxl.load_workbook(loc)
sheet = wb.active

wait = 2

driver = webdriver.Chrome('./chromedriver')

for row in sheet.iter_rows(max_col=2):
    try:
        if row[1].value:
            pass
        else:
            print('waith' + str(wait))
            number = str(row[0].value)
            driver.get('https://web.whatsapp.com/send?phone=+91' + number + '&text=' + message)

            WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')))
            element = driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div[3]')
            element.click()
            # messages = driver.find_elements_by_class_name('Txt2p')


            time.sleep(wait)
            if driver.find_elements_by_class_name('_2f-RV')[-1].find_elements_by_tag_name('span')[-1].get_attribute('data-icon') == 'msg-time':
                print('internet is slow')
                time.sleep(wait+3)
                if driver.find_elements_by_class_name('_2f-RV')[-1].find_elements_by_tag_name('span')[-1].get_attribute(
                    'data-icon') == 'msg-time':
                    print('Please check the internet!!!')
                    time.sleep(wait+5)
            else:
                sheet.cell(row=row[1].row, column=row[1].column).value ='Sent Successfully'
                wb.save(loc)

    except:
        print('Error at ' + number)

